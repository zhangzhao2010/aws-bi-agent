"""
Excel分析工具
读取并分析Excel文件结构，提取前N行数据供LLM判断
"""
from openpyxl import load_workbook
from typing import Dict, List, Any, Optional


class ExcelAnalyzer:
    """Excel文件分析器"""

    def __init__(self, file_path: str):
        self.file_path = file_path
        self.workbook = None

    def __enter__(self):
        self.workbook = load_workbook(self.file_path, data_only=True)
        return self

    def __exit__(self, exc_type, exc_val, exc_tb):
        if self.workbook:
            self.workbook.close()

    def analyze(self, sheet_name: Optional[str] = None, preview_rows: int = 100) -> Dict[str, Any]:
        """
        分析Excel文件结构

        Args:
            sheet_name: 指定sheet名称，None则分析所有sheet
            preview_rows: 预览前N行数据

        Returns:
            分析结果字典
        """
        if not self.workbook:
            raise RuntimeError("需要在context manager中使用")

        result = {
            "sheets": self.workbook.sheetnames,
            "sheet_data": {}
        }

        # 决定要分析哪些sheet
        sheets_to_analyze = [sheet_name] if sheet_name else self.workbook.sheetnames

        for sheet in sheets_to_analyze:
            if sheet not in self.workbook.sheetnames:
                continue

            ws = self.workbook[sheet]
            sheet_analysis = self._analyze_sheet(ws, preview_rows)
            result["sheet_data"][sheet] = sheet_analysis

        return result

    def _analyze_sheet(self, worksheet, preview_rows: int) -> Dict[str, Any]:
        """
        分析单个sheet

        Args:
            worksheet: openpyxl worksheet对象
            preview_rows: 预览行数

        Returns:
            sheet分析结果
        """
        # 获取sheet的实际使用范围
        max_row = worksheet.max_row
        max_col = worksheet.max_column

        # 读取前N行数据
        preview_data = []
        actual_preview_rows = min(preview_rows, max_row)

        for row_idx in range(1, actual_preview_rows + 1):
            row_data = []
            for col_idx in range(1, max_col + 1):
                cell = worksheet.cell(row_idx, col_idx)
                cell_value = cell.value

                # 转换为可序列化的格式
                if cell_value is None:
                    cell_value = ""
                elif not isinstance(cell_value, (str, int, float, bool)):
                    cell_value = str(cell_value)

                row_data.append({
                    "column": self._get_column_letter(col_idx),
                    "value": cell_value,
                    "data_type": type(cell.value).__name__ if cell.value is not None else "NoneType"
                })
            preview_data.append(row_data)

        return {
            "total_rows": max_row,
            "total_columns": max_col,
            "preview_rows": preview_data,
            "dimensions": f"{self._get_column_letter(1)}1:{self._get_column_letter(max_col)}{max_row}"
        }

    @staticmethod
    def _get_column_letter(col_idx: int) -> str:
        """将列索引转换为字母（1->A, 2->B, ...）"""
        from openpyxl.utils import get_column_letter
        return get_column_letter(col_idx)


def analyze_excel_file(file_path: str, sheet_name: Optional[str] = None, preview_rows: int = 100) -> Dict[str, Any]:
    """
    分析Excel文件的便捷函数

    Args:
        file_path: Excel文件路径
        sheet_name: 可选的sheet名称
        preview_rows: 预览行数

    Returns:
        分析结果
    """
    with ExcelAnalyzer(file_path) as analyzer:
        return analyzer.analyze(sheet_name, preview_rows)
