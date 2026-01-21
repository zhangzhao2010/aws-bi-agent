"""
色阶应用工具 - Strands Agent Tool
"""
from strands import tool, ToolContext
from openpyxl import load_workbook
from openpyxl.formatting.rule import ColorScaleRule
from pathlib import Path
import re


# 色阶配置方案
COLOR_SCHEMES = {
    "two_color": {
        "red_green": {
            "start_type": "min",
            "start_color": "F8696B",  # 红色
            "end_type": "max",
            "end_color": "63BE7B"  # 绿色
        },
        "green_red": {
            "start_type": "min",
            "start_color": "63BE7B",  # 绿色
            "end_type": "max",
            "end_color": "F8696B"  # 红色
        },
        "blue_white_red": {
            "start_type": "min",
            "start_color": "5A8AC6",  # 蓝色
            "end_type": "max",
            "end_color": "F8696B"  # 红色
        }
    },
    "three_color": {
        "red_yellow_green": {
            "start_type": "min",
            "start_color": "F8696B",  # 红色
            "mid_type": "percentile",
            "mid_value": 50,
            "mid_color": "FFEB84",  # 黄色
            "end_type": "max",
            "end_color": "63BE7B"  # 绿色
        },
        "green_yellow_red": {
            "start_type": "min",
            "start_color": "63BE7B",  # 绿色
            "mid_type": "percentile",
            "mid_value": 50,
            "mid_color": "FFEB84",  # 黄色
            "end_type": "max",
            "end_color": "F8696B"  # 红色
        },
        "blue_white_red": {
            "start_type": "min",
            "start_color": "5A8AC6",  # 蓝色
            "mid_type": "percentile",
            "mid_value": 50,
            "mid_color": "FFFFFF",  # 白色
            "end_type": "max",
            "end_color": "F8696B"  # 红色
        }
    }
}


def _calculate_cell_count(cell_range: str) -> int:
    """计算单元格范围包含的单元格数量"""
    try:
        # 解析范围如 "B2:E10"
        match = re.match(r'([A-Z]+)(\d+):([A-Z]+)(\d+)', cell_range)
        if not match:
            return 0

        def col_to_num(col):
            """列字母转数字 A->1, B->2, ..., Z->26, AA->27"""
            num = 0
            for c in col:
                num = num * 26 + (ord(c) - ord('A') + 1)
            return num

        start_col, start_row, end_col, end_row = match.groups()
        cols = col_to_num(end_col) - col_to_num(start_col) + 1
        rows = int(end_row) - int(start_row) + 1
        return cols * rows
    except:
        return 0


@tool(context=True)
def apply_color_scale(
    sheet_name: str,
    cell_range: str,
    scale_type: str,
    color_scheme: str,
    file_path: str = "",
    tool_context: ToolContext = None
) -> dict:
    """为Excel文件的指定范围应用色阶条件格式。

    此工具会：
    1. 加载指定的Excel文件
    2. 在指定的sheet和单元格范围应用色阶
    3. 保存为新文件（文件名后缀_colored）
    4. 返回新文件路径

    使用建议：
    - 先使用 analyze_excel 工具分析文件结构
    - 根据分析结果确定数据范围（跳过表头）
    - 数值数据适合应用色阶，文本数据不适合

    重要：file_path参数可以省略，系统会自动使用用户已上传的文件。

    Args:
        sheet_name: Sheet名称（如 "Sheet1"）
        cell_range: 单元格范围，格式如 "B2:E10"（注意要跳过表头）
        scale_type: 色阶类型，"two_color"（双色渐变）或 "three_color"（三色渐变：低-中-高）
        color_scheme: 色彩方案。two_color方案: "red_green"（红→绿）, "green_red"（绿→红）；three_color方案: "red_yellow_green"（红→黄→绿）, "green_yellow_red"（绿→黄→红）
        file_path: Excel文件完整路径（可选，默认使用已上传的文件）

    Returns:
        执行结果字典，格式如：
        {
          "success": true,
          "output_file": "/path/to/file_colored.xlsx",
          "sheet_name": "Sheet1",
          "applied_range": "B2:E50",
          "affected_cells": 196,
          "message": "已为 Sheet1 的 B2:E50 区域（196个单元格）应用three_color色阶"
        }
    """
    try:
        # 如果没有提供file_path，尝试从invocation_state获取
        actual_file_path = file_path
        if not file_path or file_path == "":
            if tool_context and tool_context.invocation_state:
                uploaded_files = tool_context.invocation_state.get("uploaded_files", {})
                if uploaded_files:
                    # 使用第一个上传的文件
                    actual_file_path = list(uploaded_files.values())[0]
                else:
                    return {
                        "success": False,
                        "error": "没有找到已上传的文件，请先上传Excel文件"
                    }
            else:
                return {
                    "success": False,
                    "error": "请提供file_path参数或先上传文件"
                }

        # 验证参数
        if scale_type not in COLOR_SCHEMES:
            return {
                "success": False,
                "error": f"不支持的色阶类型: {scale_type}，支持的类型: {list(COLOR_SCHEMES.keys())}"
            }

        if color_scheme not in COLOR_SCHEMES[scale_type]:
            return {
                "success": False,
                "error": f"不支持的色彩方案: {color_scheme}，支持的方案: {list(COLOR_SCHEMES[scale_type].keys())}"
            }

        # 加载工作簿
        wb = load_workbook(actual_file_path)

        # 检查sheet是否存在
        if sheet_name not in wb.sheetnames:
            return {
                "success": False,
                "error": f"Sheet '{sheet_name}' 不存在。可用的sheet: {', '.join(wb.sheetnames)}"
            }

        ws = wb[sheet_name]

        # 获取色阶配置
        scheme_config = COLOR_SCHEMES[scale_type][color_scheme]

        # 创建色阶规则
        color_scale_rule = ColorScaleRule(**scheme_config)

        # 应用色阶
        ws.conditional_formatting.add(cell_range, color_scale_rule)

        # 生成输出文件名
        path = Path(actual_file_path)
        output_file = str(path.parent / f"{path.stem}_colored{path.suffix}")

        # 保存到新文件
        wb.save(output_file)
        wb.close()

        # 计算影响的单元格数量
        cell_count = _calculate_cell_count(cell_range)

        return {
            "success": True,
            "output_file": output_file,
            "sheet_name": sheet_name,
            "applied_range": cell_range,
            "affected_cells": cell_count,
            "scale_type": scale_type,
            "color_scheme": color_scheme,
            "message": f"已为 {sheet_name} 的 {cell_range} 区域（{cell_count}个单元格）应用{scale_type}色阶"
        }

    except FileNotFoundError:
        return {
            "success": False,
            "error": f"文件不存在: {actual_file_path if 'actual_file_path' in locals() else file_path}"
        }
    except Exception as e:
        return {
            "success": False,
            "error": f"应用色阶失败: {str(e)}"
        }
