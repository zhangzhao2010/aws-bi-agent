#!/usr/bin/env python3
"""
Excel 色阶添加脚本
功能：为指定Excel文件的sheet1中指定范围的单元格添加色阶
"""

import argparse
import sys
from openpyxl import load_workbook
from openpyxl.formatting.rule import ColorScaleRule


def add_color_scale(file_path, cell_range):
    """
    为Excel文件添加色阶

    Args:
        file_path: Excel文件路径
        cell_range: 单元格范围，格式如 "A1:D10"
    """
    try:
        # 加载工作簿
        print(f"正在加载文件: {file_path}")
        wb = load_workbook(file_path)

        # 检查是否有sheet1
        if 'Sheet1' not in wb.sheetnames:
            print(f"错误: 文件中没有找到 'Sheet1'")
            print(f"可用的sheet: {', '.join(wb.sheetnames)}")
            sys.exit(1)

        ws = wb['Sheet1']

        # 创建三色色阶规则（红-黄-绿，从低到高）
        # 默认配置：
        # - 最小值：红色 (F8696B)
        # - 中间值：黄色 (FFEB84)
        # - 最大值：绿色 (63BE7B)
        color_scale_rule = ColorScaleRule(
            start_type='min',
            start_color='F8696B',  # 红色
            mid_type='percentile',
            mid_value=50,
            mid_color='FFEB84',    # 黄色
            end_type='max',
            end_color='63BE7B'     # 绿色
        )

        # 应用色阶规则到指定范围
        print(f"正在为范围 {cell_range} 添加色阶...")
        ws.conditional_formatting.add(cell_range, color_scale_rule)

        # 保存文件（覆盖原文件）
        print(f"正在保存文件...")
        wb.save(file_path)

        print(f"✓ 完成！已为 {cell_range} 添加色阶并保存")

    except FileNotFoundError:
        print(f"错误: 找不到文件 '{file_path}'")
        sys.exit(1)
    except Exception as e:
        print(f"错误: {str(e)}")
        sys.exit(1)


def main():
    parser = argparse.ArgumentParser(
        description='为Excel文件的Sheet1添加色阶',
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
示例:
  python add_color_scale.py Spend_Chart_Daily.xlsx B2:E10
  python add_color_scale.py data.xlsx A1:Z100
        """
    )

    parser.add_argument('file', help='Excel文件路径')
    parser.add_argument('range', help='单元格范围，格式如 A1:D10')

    args = parser.parse_args()

    add_color_scale(args.file, args.range)


if __name__ == '__main__':
    main()
